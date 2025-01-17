import openpyxl
from django.core.management.base import BaseCommand

from data.models import (
    Event,
    Pairing,
    List,
    Participant,
    Player,
    SNL,
)


# Function to parse the snl_results.xlsx file and build a Python dictionary
def parse_snl_results_file(filename):
    # Load the Excel file
    wb = openpyxl.load_workbook(filename)

    # Get the workbook sheets
    tournaments_sheet = wb["Tournaments"]
    players_sheet = wb["Players"]
    pairings_sheet = wb["Pairings"]

    # Iterate through the tournaments sheet and build a list of tournaments
    tournaments = []
    for row in tournaments_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        name = row[1].value
        attendee_count = int(row[2].value)
        start_date = row[3].value
        country = row[4].value
        state = row[5].value
        location = row[6].value
        max_players = int(row[7].value)

        tournament = {
            "tournamentID": tournament_id,
            "name": name,
            "attendeeCount": attendee_count,
            "startDate": start_date,
            "country": country,
            "state": state,
            "location": location,
            "maxPlayers": max_players,
        }

        tournaments.append(tournament)

    # Iterate through the players sheet and build a dictionary of players
    players = {}
    for row in players_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        registration_id = str(row[1].value)
        player_name = row[2].value
        faction = row[3].value
        army = row[4].value
        club = row[5].value
        list_plain_text = row[6].value

        player = {
            "tournamentID": tournament_id,
            "registrationID": registration_id,
            "playerName": player_name,
            "faction": faction,
            "army": army,
            "club": club,
            "listPlainText": list_plain_text,
        }

        if tournament_id in players:
            players[tournament_id].append(player)
        else:
            players[tournament_id] = [player]

    # Iterate through the pairings sheet and build a dictionary of pairings
    pairings = {}
    for row in pairings_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        round_number = int(row[1].value)
        player_id1 = str(row[2].value)
        player_result1 = row[3].value
        player_score1 = int(row[4].value)
        player_bonus1 = int(row[5].value)
        player_id2 = str(row[6].value)
        player_result2 = row[7].value
        player_score2 = int(row[8].value)
        player_bonus2 = int(row[9].value)

        pairing = {
            "tournamentID": tournament_id,
            "roundNumber": round_number,
            "playerId1": player_id1,
            "playerResult1": player_result1,
            "playerScore1": player_score1,
            "playerBonus1": player_bonus1,
            "playerId2": player_id2,
            "playerResult2": player_result2,
            "playerScore2": player_score2,
            "playerBonus2": player_bonus2,
        }

        if tournament_id in pairings:
            pairings[tournament_id].append(pairing)
        else:
            pairings[tournament_id] = [pairing]

    # Return the parsed data
    return tournaments, players, pairings


def store_tournaments(tournaments):
    for tournament in tournaments:
        source = SNL
        source_id = tournament["tournamentID"]
        source_json = tournament

        if Event.objects.filter(source=source, source_id=source_id).exists():
            continue
        event, _ = Event.objects.update_or_create(
            name=tournament["name"],
            source=source,
            source_id=source_id,
            defaults={
                "source_json": source_json,
                "start_date": tournament["startDate"],
                "players_count": tournament["attendeeCount"],
                "points_limit": 2000,
                "rounds": 5,
            }
        )
        print(f"Stored tournament {event.name}")


def store_players(players, tournament_id):
    for player in players:
        source = SNL
        source_id = player["registrationID"]
        source_json = player

        if not Player.objects.filter(source=source, source_id=source_id).exists():
            player_instance, _ = Player.objects.update_or_create(
                source=source, source_id=source_id, defaults={"source_json": source_json}
            )
        else:
            player_instance = Player.objects.get(source=source, source_id=source_id)

        participant, _ = Participant.objects.get_or_create(
            event=Event.objects.get(source_id=tournament_id),
            player=player_instance,
            defaults={"source_json": player},
        )

        player_list, _ = List.objects.get_or_create(
            participant=participant,
            source_id=player["registrationID"],
            defaults={
                "source_json": player,
                "raw_list": str(player["listPlainText"]),
            }
        )
        print(f"Stored player {player['playerName']}")


def store_pairings(pairings, tournament_id):
    for pairing in pairings:
        source = SNL
        # Generate a unique source_id for the pairing
        source_id = f"{tournament_id}_{pairing['roundNumber']}_{pairing['playerId1']}_{pairing['playerId2']}"
        source_json = pairing

        try:
            player1 = Player.objects.get(source=SNL, source_id=pairing["playerId1"])
            participant1 = Participant.objects.get(
                event__source_id=tournament_id, player=player1
            )
            player2 = Player.objects.get(source=SNL, source_id=pairing["playerId2"])
            participant2 = Participant.objects.get(
                event__source_id=tournament_id, player=player2
            )
        except Exception as e:
            print(
                f"Failed to store pairings for {tournament_id} error: {e} player1: {pairing['playerId1']} player2: {pairing['playerId2']}"
            )
            continue
        player1_list = List.objects.get(participant__player=player1)
        player2_list = List.objects.get(participant__player=player2)

        # Determine winner, loser, and is_draw
        if pairing["playerResult1"] == "WIN":
            is_draw = False
            winner = participant1
            loser = participant2
            winner_list = player1_list
            loser_list = player2_list
            winner_score = pairing["playerScore1"]
            loser_score = pairing["playerScore2"]
        elif pairing["playerResult2"] == "WIN":
            is_draw = False
            winner = participant2
            loser = participant1
            winner_list = player2_list
            loser_list = player1_list
            winner_score = pairing["playerScore2"]
            loser_score = pairing["playerScore1"]
        else:
            is_draw = True
            winner = None
            loser = None
            winner_list = None
            loser_list = None
            winner_score = pairing["playerScore1"]
            loser_score = pairing["playerScore2"]

        if not Pairing.objects.filter(
            source_id=source_id,
        ).exists():
            pairing = Pairing.objects.create(
                event=Event.objects.get(source_id=tournament_id),
                round=pairing["roundNumber"],
                source_id=source_id,
                source_json=source_json,
                winner=winner,
                loser=loser,
                is_draw=is_draw,
                winner_list=winner_list,
                loser_list=loser_list,
                winner_score=winner_score,
                loser_score=loser_score,
            )


def store_data(filename):
    tournaments, players, pairings = parse_snl_results_file(filename)

    store_tournaments(tournaments)
    for tournament_id, player_data in players.items():
        store_players(player_data, tournament_id)

    for tournament_id, pairing_data in pairings.items():
        try:
            print(f"Storing pairings for {tournament_id} pairings: {len(pairing_data)}")
            store_pairings(pairing_data, tournament_id)
            print(
                f"Stored pairings for {tournament_id}, total pairings: {Pairing.objects.filter(event__source_id=tournament_id).count()}"
            )
        except Exception as e:
            print(f"Failed to store pairings for {tournament_id} error: {e}")
            continue


class Command(BaseCommand):
    help = "Fetch events from SNL"

    def handle(self, *args, **options):
        filename = "snl_results.xlsx"
        store_data(filename)
